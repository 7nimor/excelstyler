import pytest
import openpyxl
from openpyxl import Workbook
from excelstyler.values import create_value


class TestCreateValue:
    """Test cases for create_value function."""
    
    def setup_method(self):
        """Set up test workbook and worksheet."""
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
    
    def test_create_value_basic(self):
        """Test basic value creation."""
        data = ["John", 25, "New York"]
        create_value(self.worksheet, data, 1, 1)
        
        assert self.worksheet.cell(1, 1).value == "John"
        assert self.worksheet.cell(1, 2).value == 25
        assert self.worksheet.cell(1, 3).value == "New York"
    
    def test_create_value_with_number_formatting(self):
        """Test value creation with number formatting."""
        data = [1000, 2500, 0]
        create_value(self.worksheet, data, 1, 1)
        
        # Check number formatting for non-zero numbers
        assert self.worksheet.cell(1, 1).number_format == '#,###'
        assert self.worksheet.cell(1, 2).number_format == '#,###'
        # Zero should not have number formatting
        assert self.worksheet.cell(1, 3).number_format == 'General'
    
    def test_create_value_with_border(self):
        """Test value creation with border."""
        data = ["John", 25]
        create_value(self.worksheet, data, 1, 1, border_style="thin")
        
        # Check that cells have borders
        assert self.worksheet.cell(1, 1).border.left.style == "thin"
        assert self.worksheet.cell(1, 2).border.left.style == "thin"
    
    def test_create_value_with_alternating_colors(self):
        """Test value creation with alternating row colors."""
        data = ["John", 25]
        create_value(self.worksheet, data, 1, 1, m=2)  # m=2 means even, should get light fill
        
        # Check that cells have light fill
        assert self.worksheet.cell(1, 1).fill.start_color.index == "00FAF0E7"
        assert self.worksheet.cell(1, 2).fill.start_color.index == "00FAF0E7"
    
    def test_create_value_with_color(self):
        """Test value creation with specific color."""
        data = ["John", 25]
        create_value(self.worksheet, data, 1, 1, color="green")
        
        # Check that cells have green fill
        assert self.worksheet.cell(1, 1).fill.start_color.index == "0000B050"
        assert self.worksheet.cell(1, 2).fill.start_color.index == "0000B050"
    
    def test_create_value_with_different_cell_highlighting(self):
        """Test value creation with different cell highlighting."""
        data = ["John", 25, "Active"]
        create_value(self.worksheet, data, 1, 1, different_cell=1, different_value=25)
        
        # Check that the cell with value 25 has red fill
        assert self.worksheet.cell(1, 2).fill.start_color.index == "00FCDFDC"
    
    def test_create_value_with_item_color(self):
        """Test value creation with specific item color."""
        data = ["John", 25, "New York"]
        create_value(self.worksheet, data, 1, 1, item_num=1, item_color="red")
        
        # Check that the second item (index 1) has red fill
        assert self.worksheet.cell(1, 2).fill.start_color.index == "00FCDFDC"
    
    def test_create_value_with_height_width(self):
        """Test value creation with height and width."""
        data = ["John", 25]
        create_value(self.worksheet, data, 1, 1, height=30, width=15)
        
        assert self.worksheet.row_dimensions[2].height == 30  # start_col + 1
        assert self.worksheet.column_dimensions['A'].width == 15
        assert self.worksheet.column_dimensions['B'].width == 15
