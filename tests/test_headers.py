import pytest
import openpyxl
from openpyxl import Workbook
from excelstyler.headers import create_header, create_header_freez


class TestCreateHeader:
    """Test cases for create_header function."""
    
    def setup_method(self):
        """Set up test workbook and worksheet."""
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
    
    def test_create_header_basic(self):
        """Test basic header creation."""
        data = ["Name", "Age", "City"]
        create_header(self.worksheet, data, 1, 1)
        
        assert self.worksheet.cell(1, 1).value == "Name"
        assert self.worksheet.cell(1, 2).value == "Age"
        assert self.worksheet.cell(1, 3).value == "City"
    
    def test_create_header_with_color(self):
        """Test header creation with color."""
        data = ["Name", "Age"]
        create_header(self.worksheet, data, 1, 1, color="green")
        
        # Check that cells have green fill
        assert self.worksheet.cell(1, 1).fill.start_color.index == "0000B050"
        assert self.worksheet.cell(1, 2).fill.start_color.index == "0000B050"
    
    def test_create_header_with_height_width(self):
        """Test header creation with height and width."""
        data = ["Name", "Age"]
        create_header(self.worksheet, data, 1, 1, height=30, width=15)
        
        assert self.worksheet.row_dimensions[1].height == 30
        assert self.worksheet.column_dimensions['A'].width == 15
        assert self.worksheet.column_dimensions['B'].width == 15
    
    def test_create_header_with_border(self):
        """Test header creation with border."""
        data = ["Name", "Age"]
        create_header(self.worksheet, data, 1, 1, border_style="thin")
        
        # Check that cells have borders
        assert self.worksheet.cell(1, 1).border.left.style == "thin"
        assert self.worksheet.cell(1, 2).border.left.style == "thin"
    
    def test_create_header_none_worksheet(self):
        """Test create_header with None worksheet raises ValueError."""
        with pytest.raises(ValueError, match="Worksheet cannot be None"):
            create_header(None, ["Name"], 1, 1)
    
    def test_create_header_none_data(self):
        """Test create_header with None data raises ValueError."""
        with pytest.raises(ValueError, match="Data must be a non-empty list"):
            create_header(self.worksheet, None, 1, 1)
    
    def test_create_header_empty_data(self):
        """Test create_header with empty data raises ValueError."""
        with pytest.raises(ValueError, match="Data must be a non-empty list"):
            create_header(self.worksheet, [], 1, 1)
    
    def test_create_header_invalid_start_col(self):
        """Test create_header with invalid start_col raises ValueError."""
        with pytest.raises(ValueError, match="start_col and row must be positive integers"):
            create_header(self.worksheet, ["Name"], 0, 1)
    
    def test_create_header_invalid_row(self):
        """Test create_header with invalid row raises ValueError."""
        with pytest.raises(ValueError, match="start_col and row must be positive integers"):
            create_header(self.worksheet, ["Name"], 1, 0)


class TestCreateHeaderFreez:
    """Test cases for create_header_freez function."""
    
    def setup_method(self):
        """Set up test workbook and worksheet."""
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
    
    def test_create_header_freez_basic(self):
        """Test basic frozen header creation."""
        data = ["Name", "Age", "City"]
        create_header_freez(self.worksheet, data, 1, 2, 3)
        
        assert self.worksheet.cell(2, 1).value == "Name"
        assert self.worksheet.cell(2, 2).value == "Age"
        assert self.worksheet.cell(2, 3).value == "City"
        
        # Check freeze panes
        assert self.worksheet.freeze_panes == "A3"
    
    def test_create_header_freez_with_auto_filter(self):
        """Test frozen header with auto filter."""
        data = ["Name", "Age"]
        create_header_freez(self.worksheet, data, 1, 2, 3)
        
        # Check auto filter is applied
        assert self.worksheet.auto_filter.ref is not None
    
    def test_create_header_freez_with_different_cell(self):
        """Test frozen header with different cell highlighting."""
        data = ["Name", "Age", "Status"]
        create_header_freez(self.worksheet, data, 1, 2, 3, different_cell="Status")
        
        # Check that Status cell has red fill
        assert self.worksheet.cell(2, 3).fill.start_color.index == "00C00000"