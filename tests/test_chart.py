import pytest
import openpyxl
from openpyxl import Workbook
from excelstyler.chart import add_chart


class TestAddChart:
    """Test cases for add_chart function."""
    
    def setup_method(self):
        """Set up test workbook and worksheet with sample data."""
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        
        # Add sample data
        self.worksheet.cell(1, 1, "Name")
        self.worksheet.cell(1, 2, "Value")
        self.worksheet.cell(2, 1, "Item1")
        self.worksheet.cell(2, 2, 100)
        self.worksheet.cell(3, 1, "Item2")
        self.worksheet.cell(3, 2, 200)
        self.worksheet.cell(4, 1, "Item3")
        self.worksheet.cell(4, 2, 150)
    
    def test_add_line_chart(self):
        """Test adding a line chart."""
        add_chart(
            worksheet=self.worksheet,
            chart_type='line',
            data_columns=2,
            category_column=1,
            start_row=2,
            end_row=4,
            chart_position="A6",
            chart_title="Test Line Chart",
            x_axis_title="Items",
            y_axis_title="Values"
        )
        
        # Check that chart was added
        assert len(self.worksheet._charts) == 1
        chart = self.worksheet._charts[0]
        assert chart.title.tx.rich.p[0].r[0].t == "Test Line Chart"
        assert chart.y_axis.title.tx.rich.p[0].r[0].t == "Values"
        assert chart.x_axis.title.tx.rich.p[0].r[0].t == "Items"
    
    def test_add_bar_chart(self):
        """Test adding a bar chart."""
        add_chart(
            worksheet=self.worksheet,
            chart_type='bar',
            data_columns=2,
            category_column=1,
            start_row=2,
            end_row=4,
            chart_position="A6",
            chart_title="Test Bar Chart",
            x_axis_title="Items",
            y_axis_title="Values"
        )
        
        # Check that chart was added
        assert len(self.worksheet._charts) == 1
        chart = self.worksheet._charts[0]
        assert chart.title.tx.rich.p[0].r[0].t == "Test Bar Chart"
        assert chart.y_axis.title.tx.rich.p[0].r[0].t == "Values"
        assert chart.x_axis.title.tx.rich.p[0].r[0].t == "Items"
    
    def test_add_chart_invalid_type(self):
        """Test adding chart with invalid type raises ValueError."""
        with pytest.raises(ValueError, match="chart_type must be 'line' or 'bar'"):
            add_chart(
                worksheet=self.worksheet,
                chart_type='pie',
                data_columns=2,
                category_column=1,
                start_row=2,
                end_row=4,
                chart_position="A6",
                chart_title="Test Chart",
                x_axis_title="Items",
                y_axis_title="Values"
            )
    
    def test_add_chart_with_custom_size(self):
        """Test adding chart with custom size."""
        add_chart(
            worksheet=self.worksheet,
            chart_type='line',
            data_columns=2,
            category_column=1,
            start_row=2,
            end_row=4,
            chart_position="A6",
            chart_title="Test Chart",
            x_axis_title="Items",
            y_axis_title="Values",
            chart_width=30,
            chart_height=20
        )
        
        # Check that chart was added with custom size
        assert len(self.worksheet._charts) == 1
        chart = self.worksheet._charts[0]
        assert chart.width == 30
        assert chart.height == 20
