from openpyxl.styles import PatternFill, Alignment, Font

blue_fill = PatternFill(start_color="277358", fill_type="solid")
Alignment_CELL = Alignment(horizontal='center', vertical='center', wrap_text=True)
red_font = Font(color="C00000", bold=True)
GREEN_CELL = PatternFill(start_color="00B050", fill_type="solid")
RED_CELL = PatternFill(start_color="FCDFDC", fill_type="solid")
YELLOW_CELL = PatternFill(start_color="FFFF00", fill_type="solid")
ORANGE_CELL = PatternFill(start_color="FFC000", fill_type="solid")
BLUE_CELL = PatternFill(start_color="538DD5", fill_type="solid")
LIGHT_GREEN_CELL = PatternFill(start_color="92D050", fill_type="solid")
VERY_LIGHT_GREEN_CELL = PatternFill(start_color="5AFC56", fill_type="solid")


color_dict = {
        'green': GREEN_CELL,
        'yellow': YELLOW_CELL,
        'blue': BLUE_CELL,
        'red': RED_CELL,
        'light_green': LIGHT_GREEN_CELL,
        'very_light_green': VERY_LIGHT_GREEN_CELL
    }
