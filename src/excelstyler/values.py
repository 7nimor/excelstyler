import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from .styles import *


def create_value(worksheet, data, start_col, row, border_style=None, m=None, height=None, color=None, width=None,
                 different_cell=None, different_value=None, item_num=None, item_color=None):

    for item in range(len(data)):
        cell = worksheet.cell(row=start_col, column=item + row, value=data[item])
        cell.alignment = Alignment_CELL

        if border_style:
            cell.border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style=border_style),
                right=openpyxl.styles.Side(style=border_style),
                top=openpyxl.styles.Side(style=border_style),
                bottom=openpyxl.styles.Side(style=border_style)
            )

        value = data[item]
        if isinstance(value, (int, float)) and value != 0:
            cell.number_format = '#,###'
        else:
            cell.value = value

        cell.font = Font(size=10, bold=True)

        if m is not None and m % 2 != 0:
            cell.fill = PatternFill(start_color="D6F6FE", fill_type="solid")

        if height is not None:
            worksheet.row_dimensions[start_col + 1].height = height

        if item_num is not None and item == item_num:
            if item_color:
                cell.fill = item_color
        elif color in color_dict:
            cell.fill = color_dict[color]

        if different_cell is not None and data[different_cell] == different_value:
            cell.fill = RED_CELL

        if width is not None:
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(item + row)].width = width
