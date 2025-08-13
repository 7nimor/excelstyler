import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from .styles import *


def create_header(worksheet, data, start_col, row, height=None, width=None, color=None, text_color=None, border_style=None):
    for col_num, option in enumerate(data, start_col):
        cell = worksheet.cell(row=row, column=col_num, value=option)
        col_letter = get_column_letter(col_num)
        cell.alignment = Alignment_CELL
        if color is not None:
            if color in color_dict:
                cell.fill = color_dict[color]
            else:
                cell.fill = PatternFill(start_color=color, fill_type="solid")
        else:
            cell.fill = GREEN_CELL
        if text_color is not None:
            cell.font = Font(size=9, bold=True, color=text_color)
        else:
            cell.font = Font(size=9, bold=True, color='D9FFFFFF')
        if height is not None:
            worksheet.row_dimensions[row].height = height
        if width is not None:
            worksheet.column_dimensions[col_letter].width = width
        if border_style is not None:
            cell.border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style=border_style),
                right=openpyxl.styles.Side(style=border_style),
                top=openpyxl.styles.Side(style=border_style),
                bottom=openpyxl.styles.Side(style=border_style)
            )


def create_header_freez(worksheet, data, start_col, row, header_row, height=None, width=None, len_with=None,
                        different_cell=None, color=None):
    for col_num, option in enumerate(data, start_col):
        col_letter = get_column_letter(col_num)
        cell = worksheet.cell(row=row, column=col_num, value=option)
        cell.alignment = Alignment_CELL
        cell.fill = blue_fill
        if color is not None:
            if color in color_dict:
                cell.fill = color_dict[color]
            else:
                cell.fill = PatternFill(start_color=color, fill_type="solid")
        else:
            cell.fill = GREEN_CELL

        if height is not None:
            worksheet.row_dimensions[row].height = height
            if len(option) > worksheet.column_dimensions[col_letter].width:
                worksheet.column_dimensions[col_letter].width = len(option) + 2
        if width is not None:
            worksheet.column_dimensions[col_letter].width = width
        if len_with is not None:
            if len(option) > worksheet.column_dimensions[col_letter].width:
                worksheet.column_dimensions[col_letter].width = len(option) + 3
        if different_cell is not None:
            if option == different_cell:
                cell.fill = PatternFill(start_color="C00000", fill_type="solid")
        worksheet.freeze_panes = worksheet[f'A{header_row}']
        max_col = worksheet.max_column
        range_str = f'A{header_row - 1}:{get_column_letter(max_col)}{worksheet.max_row}'
        worksheet.auto_filter.ref = range_str
